from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent


from django.contrib import admin
from .models import *

from import_export import resources
from .models import Dorminspect

from openpyxl import Workbook
from openpyxl.styles import Font,Alignment

from .views import download1,download2

admin.site.site_title = '纪检统分系统'
admin.site.site_header = '纪检统分系统后台'



# Register your models here.
@admin.register(Dorminspect)
class DorminspectAdmin(admin.ModelAdmin):
    list_display = ['id','person','room','add_date','score']
    list_filter = ['add_date']
    search_fields = ['room']
    readonly_fields = ['score']
    ordering = ['room']

    # def getdatasforteachers(self,request,queryset):
    #     temp= []
    #     for d in queryset:
    #         t = [d.room,str(d.score)]
    #         temp.append(t)
    #     f = open('D://dataforteachers.txt','a')
    #     for t in temp:
    #         f.write(','.join(t)+'\r\n')
    #     f.close
    #     self.message_user(request,'数据导出成功！')
    # getdatasforteachers.short_description = '导出提交给老师的数据'


    def getdataforteachers(self,request,queryset):
        temp = []
        for d in queryset:
            t = [d.room,str(d.score)]
            temp.append(t)
        wb = Workbook()
        ws = wb.create_sheet('mytest', 0)
        #ws.alignment=openpyxl.styles.Alignment(wrapText=True)
        start_col = 1
        haha = ['room', 'score']
        for hahaha in haha:
            ws.cell(row=1,column=start_col,value=hahaha)
            start_col += 1
        font = Font(bold=True)
        ws['A1'].font = font
        ws['B1'].font = font
        start_row = 2
        start_col = 1
        init_col = start_col
        for temp2 in temp:
            for each_temp in temp2:
                ws.cell(row=start_row, column=start_col, value=each_temp)

                start_col += 1
            start_row += 1
            start_col = init_col
        SAVE_DIR = BASE_DIR / 'index/static/给老师的数据.xlsx'
        wb.save(SAVE_DIR)
        wb.close
        self.message_user(request, '数据导出成功！')
        return download1(request)
    getdataforteachers.short_description = '导出提交给老师的数据-excel'



    def details(self,request,queryset):
        #room = ['3-301', '302', '303']

        list2 = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'B1-1', 'B1-2', 'B1-3', 'B1-4', 'B1-5', 'B1-6', 'B2', 'B3-1', 'B3-2', 'B3-3', 'B3-4', 'B3-5', 'B3-6', 'B4-1', 'B4-2', 'B4-3', 'B4-4', 'B4-5', 'B4-6', 'B5-1', 'B5-2', 'B5-3', 'B5-4', 'B5-5', 'B5-6', 'B6-1', 'B6-2', 'B6-3', 'B6-4', 'B6-5', 'B6-6', 'B7-1', 'B7-2', 'B7-3', 'B7-4', 'B7-5', 'B7-6', 'B8-1', 'B8-2', 'B8-3', 'B8-4', 'B8-5', 'B8-6', 'B9-1', 'B9-2', 'B9-3', 'B9-4', 'B9-5', 'B9-6', 'C1-1', 'C1-2', 'C1-3', 'C1-4', 'C1-5', 'C1-6', 'C2-1', 'C2-2', 'C2-3', 'C2-4', 'C2-5', 'C2-6', 'C3-1', 'C3-2', 'C3-3', 'C3-4', 'C3-5', 'C3-6', 'C4-12', 'C4-34', 'C4-56', 'C5-12', 'C5-34', 'C5-56', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D9', 'D10', 'D11', 'D12', 'D13', 'D14', 'D15', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'E13', 'E14', 'E15', 'E16', 'E17', 'E18', 'E19', 'E20', 'F1', 'F2', 'F3']
        #list1 = [True, 23, False, 0, 12]
        #list2 = ['aaaaaaa', 'bbbbbb', 'c', 'd', 'eeeeeeeeee']

        listoi = []

        for d in queryset:
            list1 = [d.ctjh, d.ctzw, d.ckjh, d.ccjh, d.scbzj, d.mjh, d.blct, d.blbzj, d.mc, d.cl, d.zszw1, d.zszw2, d.zszw3, d.zszw4, d.zszw5, d.zszw6, d.zswp, d.sjzw1, d.sjzw2, d.sjzw3, d.sjzw4, d.sjzw5, d.sjzw6, d.cmwgj1, d.cmwgj2, d.cmwgj3, d.cmwgj4, d.cmwgj5, d.cmwgj6, d.yz1, d.yz2, d.yz3, d.yz4, d.yz5, d.yz6, d.ct1, d.ct2, d.ct3, d.ct4, d.ct5, d.ct6, d.yzzw1, d.yzzw2, d.yzzw3, d.yzzw4, d.yzzw5, d.yzzw6, d.cmbzj1, d.cmbzj2, d.cmbzj3, d.cmbzj4, d.cmbzj5, d.cmbzj6, d.cmgt1, d.cmgt2, d.cmgt3, d.cmgt4, d.cmgt5, d.cmgt6, d.cszw1, d.cszw2, d.cszw3, d.cszw4, d.cszw5, d.cszw6, d.csyp1, d.csyp2, d.csyp3, d.csyp4, d.csyp5, d.csyp6, d.wz1, d.wz2, d.wz3, d.wz4, d.wz5, d.wz6, d.shoes1, d.shoes2, d.shoes3, d.cxzw1, d.cxzw2, d.cxzw3, d.dmbzj, d.wjdq, d.szbj, d.bjlj, d.szjh, d.dqwg, d.lpbzq, d.lpzw, d.lpjzw, d.qj, d.kgczjh, d.gyss, d.xlx, d.qsgt, d.wsmf, d.stbgj, d.styzw, d.szb, d.scjh, d.scsz, d.mjxg, d.ywls, d.mt, d.tb, d.ljt, d.ljtlj, d.wsjxsdk, d.wsjdmzw, d.wsjmf, d.wsjdmgz, d.ysqbbgj, d.ysdbbgj, d.qjby, d.wsjqgt, d.wsjdd, d.zlzw, d.zldd, d.zlmf]
            temp = []
            for n in range(len(list1)):
                if type(list1[n]) == bool:
                    if int(list1[n]) == 1:
                        temp.append(n)
                else:
                    if list1[n] != 0:
                        temp.append(n)
            kf = ''
            for m in temp:
                kf += list2[m] + ' '
            listoi.append([d.room, kf,d.score])

        wb = Workbook()
        ws = wb.create_sheet('detail', 0)
        start_col = 1
        haha = ['寝室', '扣分条目','分数']
        for hahaha in haha:
            ws.cell(row=1,column=start_col,value=hahaha)
            start_col += 1
        font = Font(bold=True)
        ws['A1'].font = font
        ws['B1'].font = font
        ws['C1'].font = font
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['A1'].alignment = align
        ws['B1'].alignment = align
        ws['C1'].alignment = align
        ws.column_dimensions['B'].width = 80.0
        start_row = 2
        start_col = 1
        init_col = start_col
        for temp2 in listoi:
            for each_temp in temp2:

                ws.cell(row=start_row, column=start_col, value=each_temp)
                ws.cell(row=start_row, column=start_col,).alignment = align
                start_col += 1
            start_row += 1
            start_col = init_col
        SAVE_DIR = BASE_DIR / 'index/static/详细数据.xlsx'
        wb.save(SAVE_DIR)
        wb.close
        self.message_user(request, '数据导出成功！')
        return download2(request)
    details.short_description = '详细数据-excel'


    actions = ['getdataforteachers','details']


@admin.register(PersonInfo)
class personinfoadmin(admin.ModelAdmin):
    list_display = ['id','name']